"""
Writes a job's box list into a COPY of the Dispatch Works Order template
(never the original - this produces a review copy, not a replacement for
opening the real file in Excel).

Only fills D2 / F4:H4 / F8:J8-down - it leaves every formula (K/L box
length, the P:Y sorted view, the label sheets, DISPATCH WORKS ORDER) exactly
as the template already has them, so opening this in Excel looks and
behaves identically to filling it in by hand.
"""

import shutil
import openpyxl
from pathlib import Path
from typing import List

from box_order.box_grouping import Box

PACKAGE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = PACKAGE_DIR / "template" / "TEMPLATE Dispatch Works Order Form- V DEC 20232.xlsx"
DEFAULT_OUT_DIR = PACKAGE_DIR / "output"

FIRST_BOX_ROW = 8
LAST_TEMPLATE_ROW = 117  # matches the template's 110 pre-built box rows


def write_job(
    job_id: str,
    client: str,
    boxes: List[Box],
    total_hoods: int,
    total_h_sections: int,
    total_joiners: int,
    out_name: str,
    template: Path = DEFAULT_TEMPLATE,
    out_dir: Path = DEFAULT_OUT_DIR,
) -> Path:
    if len(boxes) > (LAST_TEMPLATE_ROW - FIRST_BOX_ROW + 1):
        raise ValueError(
            f"{job_id}: {len(boxes)} boxes exceeds the template's "
            f"{LAST_TEMPLATE_ROW - FIRST_BOX_ROW + 1}-box single sheet - "
            "per the guide, save a second DWO template and continue box "
            "numbers from 21 (or the appropriate next box number)."
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / out_name
    shutil.copy(template, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["BOX ORDER"]

    ws["D2"] = f"{job_id} - {client}"
    ws["F4"] = total_hoods
    ws["G4"] = total_h_sections
    ws["H4"] = total_joiners

    for i, box in enumerate(boxes):
        row = FIRST_BOX_ROW + i
        sorted_pieces = box.sorted_pieces
        ws.cell(row=row, column=6, value=box.label)                       # F: hood label(s)
        ws.cell(row=row, column=7, value=box.depth_mm)                    # G: depth
        ws.cell(row=row, column=8, value=sorted_pieces[0].length_mm)      # H: length H1
        ws.cell(row=row, column=9,                                       # I: length H2
                 value=sorted_pieces[1].length_mm if len(sorted_pieces) > 1 else 0)
        ws.cell(row=row, column=10, value=box.returns)                    # J: returns
        if box.reasons:
            ws.cell(row=row, column=13, value="; ".join(box.reasons))     # M: notes
        ws.row_dimensions[row].hidden = False

    # hide the remaining unused template rows, same as the manual step
    last_used_row = FIRST_BOX_ROW + len(boxes) - 1
    for row in range(last_used_row + 1, LAST_TEMPLATE_ROW + 1):
        ws.row_dimensions[row].hidden = True

    wb.save(out_path)
    return out_path
